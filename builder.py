"""
builder.py
Crea o actualiza el archivo "Nuevas OC - Control de Ordenes de Compra.xlsx".

Reglas clave (ver README):
- Si el archivo de salida ya existe, se agregan filas nuevas al final;
  las columnas que el usuario llena a mano (Back order, Categoria,
  Mes de facturacion, Status) NUNCA se tocan en filas ya existentes.
- Las columnas calculadas (Precio coincide, Lista de precio, Limite de
  Credito, Dentro de Limite, Inv Disponible, Caducidad) se recalculan
  para TODAS las filas cada vez que corre el programa, porque dependen
  de datos que cambian con el tiempo.
"""

from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_HEADERS, MANUAL_COLUMNS, CUSTOMER_OVERRIDE_BY_PO_NUMBER

HEADER_ROW = 4
FIRST_DATA_ROW = 5

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="000000")
HEADER_FONT = Font(name=FONT_NAME, size=9, bold=True, color="FFFFFF")
DATA_FONT = Font(name=FONT_NAME, size=9)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
RED_FONT = Font(name=FONT_NAME, size=9, color="990000", bold=True)
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")
GREEN_FONT = Font(name=FONT_NAME, size=9, color="38761D", bold=True)


def _quarter_formula(row):
    r = row
    return (
        f'=IF(O{r}="Ene","1Q",IF(O{r}="feb","1Q",IF(O{r}="Mar","1Q",IF(O{r}="AbR","2Q",'
        f'IF(O{r}="MAY","2Q",IF(O{r}="JUN","2Q",IF(O{r}="JUL","3Q",IF(O{r}="Ago","3Q",'
        f'IF(O{r}="SEP","3Q",IF(O{r}="OCT","4Q",IF(O{r}="NOV","4Q",IF(O{r}="DiC","4Q",0))))))))))))'
    )


def _semester_formula(row):
    r = row
    return (
        f'=IF(O{r}="Ene","1S",IF(O{r}="feb","1S",IF(O{r}="Mar","1S",IF(O{r}="Abr","1S",'
        f'IF(O{r}="MAY","1S",IF(O{r}="JUN","1S",IF(O{r}="jul","2s",IF(O{r}="Ago","2S",'
        f'IF(O{r}="SEP","2S",IF(O{r}="OCT","2S",IF(O{r}="NOV","2S",IF(O{r}="DiC","2S",0))))))))))))'
    )


def _open_or_create(output_path: Path):
    if output_path.exists():
        wb = openpyxl.load_workbook(output_path, data_only=False)
        ws = wb["Purchase Order"] if "Purchase Order" in wb.sheetnames else wb.active
        return wb, ws, True
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    for i, h in enumerate(OUTPUT_HEADERS, start=1):
        c = ws.cell(row=HEADER_ROW, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return wb, ws, False


def _existing_po_numbers(ws):
    numbers = set()
    r = FIRST_DATA_ROW
    while ws.cell(row=r, column=4).value is not None or ws.cell(row=r, column=6).value is not None:
        v = ws.cell(row=r, column=4).value
        if v is not None:
            numbers.add(str(v).strip())
        r += 1
    return numbers, r  # r = primera fila libre


def build_or_update(output_path: Path, new_pos, master, log=print):
    """
    new_pos: lista de tuplas (ExtractedPO, [(line_item, matched_fields_dict), ...])
             ya resueltas contra el customer master (ver main.py).
    """
    wb, ws, existed = _open_or_create(output_path)
    existing_numbers, next_row = _existing_po_numbers(ws) if existed else (set(), FIRST_DATA_ROW)

    added = 0
    for extracted, resolved_items in new_pos:
        if extracted.po_number.strip() in existing_numbers:
            log(f"  OC {extracted.po_number} ya estaba en el archivo, se omite")
            continue
        for item, fields in resolved_items:
            r = next_row
            ws.cell(row=r, column=1, value=fields["pic"])
            ws.cell(row=r, column=2, value=fields["po_date"])
            ws.cell(row=r, column=3, value=fields["po_date"])
            ws.cell(row=r, column=4, value=extracted.po_number)
            ws.cell(row=r, column=5, value=fields["customer_code"])
            ws.cell(row=r, column=6, value=fields["customer_name"])
            ws.cell(row=r, column=7, value=item.code)
            ws.cell(row=r, column=8, value=item.description)
            ws.cell(row=r, column=9, value=fields["category"])
            ws.cell(row=r, column=10, value=item.qty)
            ws.cell(row=r, column=11, value=item.unit_price)
            # L (Back order), N (Categoria), O (Mes de facturacion), R (Status),
            # Q (Comments (SCM)) -> siempre en blanco, sin texto y sin color.
            # La info de "revisar manualmente" (cliente no encontrado, codigo
            # incierto, notas de extraccion) va solo al log/warnings (ver
            # main.py) -- la celda Q no se toca para nada mas que dejarla en
            # blanco.
            ws.cell(row=r, column=13, value=f"=J{r}*K{r}")
            ws.cell(row=r, column=19, value=_quarter_formula(r))
            ws.cell(row=r, column=20, value=_semester_formula(r))
            ws.cell(row=r, column=21, value=f"=YEAR(B{r})")
            for col in range(1, 22):
                cell = ws.cell(row=r, column=col)
                cell.font = DATA_FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(col in (8, 17)))
            next_row += 1
            added += 1
        existing_numbers.add(extracted.po_number.strip())

    last_row = next_row - 1
    if last_row >= FIRST_DATA_ROW:
        _refresh_calculated_columns(ws, FIRST_DATA_ROW, last_row, master, log)
        _apply_total_row(ws, last_row)
        _apply_autofilter_and_conditional_formatting(ws, last_row)

    wb.save(output_path)
    log(f"Guardado: {output_path} ({added} lineas nuevas agregadas)")
    return added


def _refresh_calculated_columns(ws, first_row, last_row, master, log):
    """Recalcula V..AB para TODAS las filas de datos (existentes + nuevas).
    Tambien re-intenta el cruce contra Product master list para G/H/I
    (Code/Description/Category Product) en TODAS las filas -- no solo las
    nuevas -- porque Product master list puede agregarse o actualizarse
    despues de que una fila ya fue creada, y sin este refresco esas filas
    viejas se quedarian para siempre con el codigo crudo sin limpiar.

    Tambien fuerza la columna Q (Comments (SCM)) a blanco y sin color en
    TODAS las filas, incluyendo filas viejas que hayan quedado con texto o
    fondo rojo de versiones anteriores del programa -- esta columna ya no
    lleva ningun texto ni resaltado, nunca.

    Y tambien re-intenta el cruce contra customer master list (por nombre,
    columna F) en TODAS las filas -- igual que con Product master list, esto
    corrige filas viejas si se agrega un alias nuevo despues de que la fila
    ya fue creada (ver config.CUSTOMER_NAME_ALIASES), sin tener que volver a
    subir el documento original."""
    for r in range(first_row, last_row + 1):
        po_number = ws.cell(row=r, column=4).value
        customer_name = ws.cell(row=r, column=6).value
        customer_code = ws.cell(row=r, column=5).value
        code = ws.cell(row=r, column=7).value
        description = ws.cell(row=r, column=8).value

        override_name = CUSTOMER_OVERRIDE_BY_PO_NUMBER.get(str(po_number).strip()) if po_number else None
        if override_name:
            customer_name = override_name

        customer_record = master.find_customer(name=customer_name)
        if customer_record:
            customer_code = customer_record["customer_code"]
            customer_name = customer_record["customer_name"]
            ws.cell(row=r, column=1, value=customer_record["pic"] or "No disponible")
            ws.cell(row=r, column=5, value=customer_code)
            ws.cell(row=r, column=6, value=customer_name)

        product_match = master.find_product(code, description)
        if product_match:
            code = product_match["code"]
            ws.cell(row=r, column=7, value=code)
            if product_match["description"]:
                ws.cell(row=r, column=8, value=product_match["description"])
            if product_match["category"]:
                ws.cell(row=r, column=9, value=product_match["category"])

        # OJO con el gotcha de openpyxl: asignar value=None via
        # ws.cell(..., value=None) es un no-op si la celda ya tenia
        # contenido -- hay que asignar .value directamente.
        q_cell = ws.cell(row=r, column=17)
        q_cell.value = None
        q_cell.fill = PatternFill(fill_type=None)
        q_cell.font = DATA_FONT

        unit_price = ws.cell(row=r, column=11).value or 0
        amount = ws.cell(row=r, column=13).value  # formula '=J*K'; usamos qty*price directo para comparar
        qty = ws.cell(row=r, column=10).value or 0
        amount_value = float(qty) * float(unit_price)

        customer_record = master.find_customer_by_code(customer_code)
        classification = customer_record["classification"] if customer_record else None
        rfc = customer_record["rfc"] if customer_record else None

        # V: Precio coincide
        ref_price = master.latest_price(customer_name, code)
        if ref_price is None:
            v_val = "N/D (sin historial en Sales_YTD)"
        else:
            v_val = "YES" if abs(float(ref_price) - float(unit_price)) < 0.01 else "NO"
        ws.cell(row=r, column=22, value=v_val)

        # W: Lista de precio, segun clasificacion real del cliente (customer master list)
        w_price, w_label = master.price_for_classification(code, classification)
        if w_price is None:
            ws.cell(row=r, column=23, value=w_label)
        else:
            c = ws.cell(row=r, column=23, value=w_price)
            c.number_format = "$#,##0.00"

        # X, Y: limite de credito (via RFC) y si el importe de esta linea cabe dentro del limite
        balance = master.credit_balance(rfc)
        if balance is None:
            ws.cell(row=r, column=24, value="No disponible (sin registro en CREDITO)")
            ws.cell(row=r, column=25, value="N/D")
        else:
            xc = ws.cell(row=r, column=24, value=balance)
            xc.number_format = "$#,##0.00"
            ws.cell(row=r, column=25, value="YES" if amount_value <= float(balance) else "NO")

        # Z, AA, AB: inventario y caducidad
        inv_qty, classification_exp, exp_info = master.inventory_and_expiration(code)
        if inv_qty is None:
            ws.cell(row=r, column=26, value="No disponible (codigo no encontrado en Inventory)")
        else:
            ws.cell(row=r, column=26, value=inv_qty)
        aa_cell = ws.cell(row=r, column=27, value=classification_exp)
        if isinstance(exp_info, date):
            ab_cell = ws.cell(row=r, column=28, value=exp_info)
            ab_cell.number_format = "DD/MM/YYYY"
        else:
            ws.cell(row=r, column=28, value=exp_info)

        if classification_exp == "< 6 meses":
            aa_cell.fill, aa_cell.font = RED_FILL, RED_FONT
        elif classification_exp == ">= 6 meses":
            aa_cell.fill, aa_cell.font = GREEN_FILL, GREEN_FONT

        for col in (22, 23, 24, 25, 26, 27, 28):
            cell = ws.cell(row=r, column=col)
            if col != 27 or cell.value not in ("< 6 meses", ">= 6 meses"):
                # no pisar el font rojo/verde que ya se puso arriba para AA
                cell.font = DATA_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_total_row(ws, last_row):
    # Limpia cualquier fila "TOTAL" que haya quedado de una corrida anterior
    # (si el archivo crecio, la fila de total de la vez pasada queda mas
    # arriba que la nueva y hay que borrarla para no dejarla huerfana).
    # OJO: ws.cell(row=r, column=c, value=None) NO limpia la celda en openpyxl
    # (es un no-op si la celda ya tenia contenido). Hay que asignar .value
    # directamente para que el borrado realmente se guarde en el archivo.
    for r in range(FIRST_DATA_ROW, last_row + 4):
        cell_l = ws.cell(row=r, column=12)
        if cell_l.value == "TOTAL":
            cell_l.value = None
            ws.cell(row=r, column=13).value = None

    total_row = last_row + 2
    ws.cell(row=total_row, column=12, value="TOTAL").font = Font(name=FONT_NAME, bold=True)
    c = ws.cell(row=total_row, column=13, value=f"=SUM(M{FIRST_DATA_ROW}:M{last_row})")
    c.font = Font(name=FONT_NAME, bold=True)
    c.number_format = "$#,##0.00"


def _apply_autofilter_and_conditional_formatting(ws, last_row):
    ws.auto_filter.ref = f"A{HEADER_ROW}:AB{last_row}"
    for col_letter in ("V", "Y"):
        rng = f"{col_letter}{FIRST_DATA_ROW}:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"NO"'], fill=RED_FILL, font=RED_FONT)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=['"YES"'], fill=GREEN_FILL, font=GREEN_FONT)
        )
    widths = [10, 12, 14, 16, 14, 40, 16, 40, 20, 9, 13, 12, 13, 14, 16, 24, 46, 12, 10, 10, 8,
              16, 16, 20, 14, 24, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
