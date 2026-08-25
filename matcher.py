"""
matcher.py
Cruza cada partida extraida contra los archivos maestros:
- customer master list.xlsx      -> Codigo de cliente, PIC, Clasificacion, RFC
- daily report.xlsx (Sales_YTD)  -> precio de la factura mas reciente (Precio coincide)
- daily report.xlsx (Inventory)  -> inventario disponible y caducidad
- 190626 Official sales price list.xlsx -> lista de precio segun clasificacion
- sales collection report.xlsx (CREDITO) -> limite de credito disponible

Todas las funciones son de solo lectura sobre los archivos maestros; no
modifican nada. Devuelven "None" o un texto tipo "No disponible"/"No se sabe"
cuando no hay match, nunca un valor inventado.
"""

import re
from collections import defaultdict
from datetime import date, datetime

import openpyxl

from config import (
    FILE_SALES_PROGRESS, FILE_CUSTOMER_MASTER, FILE_DAILY_REPORT,
    FILE_PRICE_LIST, FILE_CREDIT_REPORT, FILE_PRODUCT_MASTER,
    MI_SHEET, MI_HEADER_ROW, MI_FIRST_DATA_ROW, MI_COL_CODE, MI_COL_CATEGORY,
    CM_SHEET, CM_HEADER_ROW, CM_FIRST_DATA_ROW,
    CM_COL_CUSTOMER_CODE, CM_COL_CUSTOMER_NAME, CM_COL_SHORT_NAME, CM_COL_CLASSIFICATION,
    CM_COL_RFC, CM_COL_PIC, CUSTOMER_NAME_ALIASES, PRODUCT_CODE_ALIASES,
    DR_SALES_YTD_SHEET, DR_SALES_YTD_HEADER_ROW, DR_SALES_YTD_FIRST_DATA_ROW,
    DR_SYTD_COL_CUSTOMER, DR_SYTD_COL_CODE, DR_SYTD_COL_PRICE, DR_SYTD_COL_DATE,
    DR_INVENTORY_SHEET, DR_INV_HEADER_ROW, DR_INV_FIRST_DATA_ROW,
    DR_INV_COL_CATNO, DR_INV_COL_EXPIRATION, DR_INV_COL_MX_FINAL,
    PL_SHEET, PL_HEADER_ROW, PL_FIRST_DATA_ROW,
    PL_COL_CATNO, PL_COL_NEW_LIST, PL_COL_NEW_DIST_A, PL_COL_NEW_DIST_B,
    CR_SHEET, CR_HEADER_ROW, CR_FIRST_DATA_ROW,
    CR_COL_RFC, CR_COL_NAME, CR_COL_BALANCE,
    PM_SHEET, PM_HEADER_ROW, PM_FIRST_DATA_ROW,
    PM_COL_CODE, PM_COL_DESCRIPTION, PM_COL_CATEGORY,
    DISTRIBUTOR_A_LABEL, DISTRIBUTOR_B_LABEL,
)


def _optional_path(folder, filename, log, feature_desc):
    """Devuelve folder/filename si existe; si no, avisa por log y devuelve None
    en vez de dejar que openpyxl.load_workbook lance FileNotFoundError."""
    path = folder / filename
    if not path.exists():
        log(f"  AVISO: {filename} no encontrado -- {feature_desc} no estara disponible.")
        return None
    return path


def _optional_sheet(wb, sheet_name, filename, log, feature_desc):
    """Devuelve wb[sheet_name] si esa hoja existe en el archivo; si no, avisa
    por log y devuelve None en vez de dejar que falle con KeyError. El
    archivo puede existir pero tener un nombre de hoja distinto al esperado
    (plantilla vieja, hoja renombrada, etc.) -- eso no debe tronar el programa,
    solo degradar esa funcionalidad puntual."""
    if sheet_name not in wb.sheetnames:
        log(f"  AVISO: la hoja '{sheet_name}' no existe en {filename} "
            f"(hojas disponibles: {', '.join(wb.sheetnames)}) -- {feature_desc} no estara disponible.")
        return None
    return wb[sheet_name]


def _norm_header(name):
    """Normaliza un encabezado para que 'Customer \\nCode' y 'Customer Code'
    (o con espacios/saltos de linea distintos) se consideren la misma columna."""
    return " ".join(str(name).split()).strip().lower()


def _header_index(ws, header_row):
    """Devuelve {encabezado_normalizado: indice_1based} leyendo la fila de encabezado.
    Funciona tanto en modo normal como read_only (donde ws[fila] no esta disponible)."""
    idx = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row):
        for cell in row:
            if cell.value is not None:
                idx[_norm_header(cell.value)] = cell.column
        break
    return idx


def _get(row, idx, col_name):
    """Lee el valor de 'col_name' (tal como esta escrito en config.py, puede traer
    saltos de linea) en la fila dada, usando el indice normalizado de encabezados."""
    key = _norm_header(col_name)
    if key not in idx:
        return None
    return row[idx[key] - 1].value


def _norm(text):
    if text is None:
        return ""
    return str(text).strip().upper()


class MasterData:
    """Carga en memoria todos los archivos maestros una sola vez."""

    def __init__(self, folder, log=print):
        self.log = log
        self.has_customer_master = False
        self.has_sales_progress = False
        self.has_daily_report = False
        self.has_price_list = False
        self.has_credit_report = False
        self.has_product_master = False
        self._load_customer_master(folder)
        self._load_material_info(folder)
        self._load_sales_ytd(folder)
        self._load_inventory(folder)
        self._load_price_list(folder)
        self._load_credit(folder)
        self._load_product_master(folder)

        self.missing_files = []
        if not self.has_sales_progress:
            self.missing_files.append(FILE_SALES_PROGRESS)
        if not self.has_customer_master:
            self.missing_files.append(FILE_CUSTOMER_MASTER)
        if not self.has_daily_report:
            self.missing_files.append(FILE_DAILY_REPORT)
        if not self.has_price_list:
            self.missing_files.append(FILE_PRICE_LIST)
        if not self.has_credit_report:
            self.missing_files.append(FILE_CREDIT_REPORT)
        if not self.has_product_master:
            self.missing_files.append(FILE_PRODUCT_MASTER)

    # -- Material info (categoria de producto) --------------------------------
    def _load_material_info(self, folder):
        path = _optional_path(folder, FILE_SALES_PROGRESS, self.log, "categoria de producto (Material info.)")
        self.material_category = {}
        if path is None:
            return
        # read_only=True: Sales progress.xlsx tiene hojas enormes (una llega a la
        # columna XFA) -- sin read_only, cargar el libro completo puede tardar
        # varios minutos. En modo read_only openpyxl solo procesa la hoja que
        # realmente se recorre.
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = _optional_sheet(wb, MI_SHEET, FILE_SALES_PROGRESS, self.log, "categoria de producto (Material info.)")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, MI_HEADER_ROW)
        for row in ws.iter_rows(min_row=MI_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            code = get(MI_COL_CODE)
            if code is None:
                continue
            self.material_category[_norm(code)] = get(MI_COL_CATEGORY)
        wb.close()
        self.has_sales_progress = True
        self.log(f"  Material info: {len(self.material_category)} codigos cargados")

    def category_for_code(self, code):
        return self.material_category.get(_norm(code), "No disponible (no encontrado en Material info.)")

    # -- customer master list ------------------------------------------------
    def _load_customer_master(self, folder):
        self.customers_by_rfc = {}
        self.customers_by_name = {}
        self.customers_by_code = {}
        path = _optional_path(folder, FILE_CUSTOMER_MASTER, self.log, "codigo de cliente, PIC, clasificacion, RFC")
        if path is None:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = _optional_sheet(wb, CM_SHEET, FILE_CUSTOMER_MASTER, self.log, "codigo de cliente, PIC, clasificacion, RFC")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, CM_HEADER_ROW)
        for row in ws.iter_rows(min_row=CM_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            rfc = _norm(get(CM_COL_RFC))
            name = _norm(get(CM_COL_CUSTOMER_NAME))
            short_name = _norm(get(CM_COL_SHORT_NAME))
            code = get(CM_COL_CUSTOMER_CODE)
            if not name and not rfc:
                continue
            pic_raw = get(CM_COL_PIC)
            record = {
                "customer_code": code,
                "customer_name": get(CM_COL_CUSTOMER_NAME),
                "classification": get(CM_COL_CLASSIFICATION),
                "rfc": get(CM_COL_RFC),
                "pic": str(pic_raw).strip().upper() if pic_raw else pic_raw,  # PIC siempre en mayusculas en la salida
            }
            if rfc:
                self.customers_by_rfc[rfc] = record
            if name:
                self.customers_by_name[name] = record
            # El nombre corto (ej. "CONTED" para "BLANCA LIZBETH ESPINAL
            # PEREZ") se registra tambien en customers_by_name -- muchas OC
            # usan este nombre corto/comercial en vez de la razon social
            # completa, y asi se aprovechan las mismas reglas de match
            # (exacto/subcadena/palabras) sin duplicar logica. Se exige un
            # minimo de caracteres para no registrar nombres cortos
            # demasiado genericos que podrian dar falsos positivos.
            if short_name and short_name != name and len(short_name) >= 4:
                self.customers_by_name.setdefault(short_name, record)
            if code is not None:
                self.customers_by_code[code] = record
        self.has_customer_master = True
        self.log(f"  customer master list: {len(self.customers_by_rfc)} clientes con RFC cargados")

    def find_customer_by_code(self, customer_code):
        return self.customers_by_code.get(customer_code)

    _STOPWORDS = {"DE", "DEL", "LA", "LAS", "LOS", "Y", "SA", "CV", "SC", "S", "A", "C", "V"}

    @classmethod
    def _name_tokens(cls, text):
        words = re.findall(r"[A-Z0-9]+", _norm(text))
        return {w for w in words if w not in cls._STOPWORDS and len(w) > 1}

    def find_customer(self, rfc=None, name=None):
        """Busca primero por RFC (confiable), luego por alias conocido
        (ver config.CUSTOMER_NAME_ALIASES), luego por nombre (aproximado)."""
        if rfc:
            rec = self.customers_by_rfc.get(_norm(rfc))
            if rec:
                return rec
        if name:
            n = _norm(name)
            if n in self.customers_by_name:
                return self.customers_by_name[n]
            alias_target = CUSTOMER_NAME_ALIASES.get(n)
            if alias_target:
                rec = self.customers_by_name.get(_norm(alias_target))
                if rec:
                    return rec
            # coincidencia parcial simple
            for key, rec in self.customers_by_name.items():
                if key in n or n in key:
                    return rec
            # coincidencia por conjunto de palabras (sin importar el orden) --
            # Claude no siempre extrae el nombre en el mismo orden de palabras
            # entre una corrida y otra (ej. "Galindo Laboratorios" vs
            # "Laboratorios Galindo"), asi que una comparacion de texto
            # exacto o de subcadena no alcanza. Solo se acepta si hay un
            # candidato claramente mejor que el resto (nunca se adivina entre
            # dos igual de parecidos).
            n_tokens = self._name_tokens(n)
            if n_tokens:
                scored = []
                for key, rec in self.customers_by_name.items():
                    key_tokens = self._name_tokens(key)
                    if not key_tokens:
                        continue
                    overlap = len(n_tokens & key_tokens) / len(n_tokens | key_tokens)
                    if overlap >= 0.6:
                        scored.append((overlap, rec))
                if scored:
                    scored.sort(key=lambda x: -x[0])
                    if len(scored) == 1 or scored[0][0] > scored[1][0]:
                        return scored[0][1]
        return None

    # -- Sales_YTD (precio de referencia) ------------------------------------
    def _load_sales_ytd(self, folder):
        self.sales_ytd = defaultdict(list)
        path = _optional_path(folder, FILE_DAILY_REPORT, self.log, "precio de referencia (Precio coincide)")
        if path is None:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = _optional_sheet(wb, DR_SALES_YTD_SHEET, FILE_DAILY_REPORT, self.log, "precio de referencia (Precio coincide)")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, DR_SALES_YTD_HEADER_ROW)
        for row in ws.iter_rows(min_row=DR_SALES_YTD_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            cust, code, price, dt = (
                get(DR_SYTD_COL_CUSTOMER), get(DR_SYTD_COL_CODE),
                get(DR_SYTD_COL_PRICE), get(DR_SYTD_COL_DATE),
            )
            if cust is None or code is None or price is None:
                continue
            self.sales_ytd[(_norm(cust), _norm(code))].append((dt, price))
        self.has_daily_report = True
        self.log(f"  Sales_YTD: {sum(len(v) for v in self.sales_ytd.values())} renglones cargados")

    def latest_price(self, customer_name, code):
        hits = self.sales_ytd.get((_norm(customer_name), _norm(code)))
        if not hits:
            return None
        hits_sorted = sorted(hits, key=lambda x: (x[0] is None, x[0]))
        return hits_sorted[-1][1]

    # -- Inventory (existencia + caducidad) -----------------------------------
    def _load_inventory(self, folder):
        self.inventory = defaultdict(list)
        path = _optional_path(folder, FILE_DAILY_REPORT, self.log, "inventario disponible y caducidad")
        if path is None:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = _optional_sheet(wb, DR_INVENTORY_SHEET, FILE_DAILY_REPORT, self.log, "inventario disponible y caducidad")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, DR_INV_HEADER_ROW)
        for row in ws.iter_rows(min_row=DR_INV_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            catno = get(DR_INV_COL_CATNO)
            if catno is None:
                continue
            self.inventory[_norm(catno)].append({
                "expiration": get(DR_INV_COL_EXPIRATION),
                "mx_final": get(DR_INV_COL_MX_FINAL) or 0,
            })
        self.log(f"  Inventory: {len(self.inventory)} codigos distintos cargados")

    def inventory_and_expiration(self, code, today=None):
        today = today or date.today()
        lots = self.inventory.get(_norm(code))
        if not lots:
            return None, "No se sabe", "No se sabe"
        total = sum(l["mx_final"] for l in lots)
        stocked = [l for l in lots if l["mx_final"] > 0]
        if not stocked:
            return total, "No se sabe", "No se sabe"
        dated = [l for l in stocked if isinstance(l["expiration"], (date, datetime))]
        if not dated:
            return total, "N/A", "N/A"
        nearest = min(dated, key=lambda l: l["expiration"])
        exp = nearest["expiration"]
        exp_date = exp.date() if isinstance(exp, datetime) else exp
        months = (exp_date.year - today.year) * 12 + (exp_date.month - today.month) - (1 if exp_date.day < today.day else 0)
        classification = "< 6 meses" if months < 6 else ">= 6 meses"
        return total, classification, exp_date

    # -- Official price list ---------------------------------------------------
    def _load_price_list(self, folder):
        self.price_list = {}
        path = _optional_path(folder, FILE_PRICE_LIST, self.log, "lista de precio segun clasificacion")
        if path is None:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = _optional_sheet(wb, PL_SHEET, FILE_PRICE_LIST, self.log, "lista de precio segun clasificacion")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, PL_HEADER_ROW)
        for row in ws.iter_rows(min_row=PL_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            catno = get(PL_COL_CATNO)
            if catno is None:
                continue
            self.price_list[_norm(catno)] = {
                "new_list": get(PL_COL_NEW_LIST),
                "new_dist_a": get(PL_COL_NEW_DIST_A),
                "new_dist_b": get(PL_COL_NEW_DIST_B),
            }
        self.has_price_list = True
        self.log(f"  Price list: {len(self.price_list)} codigos cargados")

    def price_for_classification(self, code, classification):
        entry = self.price_list.get(_norm(code))
        if not entry:
            return None, "N/D (codigo no encontrado en Price list for contracts)"
        classification = (classification or "").strip()
        if classification == DISTRIBUTOR_A_LABEL:
            return entry["new_dist_a"], "New Price Distribuidor A"
        if classification == DISTRIBUTOR_B_LABEL:
            return entry["new_dist_b"], "New Price Distribuidor B"
        return entry["new_list"], "New List price"

    # -- CREDITO (limite de credito) --------------------------------------------
    def _load_credit(self, folder):
        self.credit_by_rfc = {}
        path = _optional_path(folder, FILE_CREDIT_REPORT, self.log, "limite de credito disponible")
        if path is None:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = _optional_sheet(wb, CR_SHEET, FILE_CREDIT_REPORT, self.log, "limite de credito disponible")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, CR_HEADER_ROW)
        for row in ws.iter_rows(min_row=CR_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            rfc = get(CR_COL_RFC)
            if rfc is None:
                continue
            self.credit_by_rfc[_norm(rfc)] = get(CR_COL_BALANCE)
        self.has_credit_report = True
        self.log(f"  CREDITO: {len(self.credit_by_rfc)} clientes con limite registrado")

    def credit_balance(self, rfc):
        if not rfc:
            return None
        return self.credit_by_rfc.get(_norm(rfc))

    # -- Product master list (codigo/descripcion/categoria de producto) ---------
    def _load_product_master(self, folder):
        self.product_by_code = {}
        path = _optional_path(folder, FILE_PRODUCT_MASTER, self.log, "codigo/descripcion/categoria de producto (Product master list)")
        if path is None:
            return
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = _optional_sheet(wb, PM_SHEET, FILE_PRODUCT_MASTER, self.log, "codigo/descripcion/categoria de producto (Product master list)")
        if ws is None:
            wb.close()
            return
        idx = _header_index(ws, PM_HEADER_ROW)
        duplicate_codes = 0
        for row in ws.iter_rows(min_row=PM_FIRST_DATA_ROW, values_only=False):
            get = lambda col: _get(row, idx, col)
            catno = get(PM_COL_CODE)
            if catno is None:
                continue
            key = _norm(catno)
            if key in self.product_by_code:
                # Codigo duplicado en Product master list (pasa con renglones
                # viejos re-agregados al final de la hoja) -- se queda con la
                # PRIMERA aparicion y se ignoran las siguientes, porque en la
                # practica las duplicadas mas abajo suelen ser texto viejo
                # /truncado (ej. sin "(100T)", con simbolos mal codificados).
                duplicate_codes += 1
                continue
            self.product_by_code[key] = {
                "code": catno,
                "description": get(PM_COL_DESCRIPTION),
                "category": get(PM_COL_CATEGORY),
            }
        wb.close()
        self.has_product_master = True
        self.log(f"  Product master list: {len(self.product_by_code)} codigos cargados"
                  + (f" ({duplicate_codes} duplicados ignorados)" if duplicate_codes else ""))

    @staticmethod
    def _tokenize(text):
        """Parte 'text' en candidatos a codigo: el texto completo primero
        (por si el codigo real de por si tiene guiones, ej. '49009-0104'),
        y luego cada pedazo separado por espacios, '/', '-', ',', ';', '*' o
        parentesis (ej. 'AD-BM-CN-SD9802X' -> 'SD9802X', o una descripcion
        larga como 'ALLPLEX STI ESSENTIALASSAY 25RX SD10245Z' -> 'SD10245Z',
        o el formato de algunos clientes 'ALLPLEX GI-VIRUS *GI9701X' -> 'GI9701X'
        entre otros pedazos)."""
        if not text:
            return []
        text = str(text)
        parts = re.split(r"[\s/,;\-()*]+", text)
        candidates = [text.strip()] + [p.strip() for p in parts if p.strip()]
        # Caso especial: cantidad pegada directamente al codigo sin separador
        # (ej. "100TB7200X" = cantidad "100" + codigo real "TB7200X"). Se
        # agrega tambien la version sin los digitos iniciales, como ultimo
        # recurso (se prueba despues de todo lo demas).
        stripped_variants = []
        for c in candidates:
            without_leading_digits = re.sub(r"^\d+", "", c)
            if without_leading_digits and without_leading_digits != c:
                stripped_variants.append(without_leading_digits)
        return candidates + stripped_variants

    def find_product(self, raw_code, description=None):
        """Busca contra Product master list, sin importar en que parte del
        texto este el codigo real. Orden de busqueda:
        1. El codigo extraido tal cual (por si ya es un codigo valido).
        2. Cada pedazo del codigo separado por espacios/guiones/slashes/etc.
           (ej. 'AD-BM-CN-SD9802X' -> prueba 'SD9802X').
        3. Si nada de lo anterior hizo match, se repite el mismo proceso
           sobre la descripcion (ej. si el codigo es 'ALG-1063486' y no
           existe en el maestro, pero la descripcion dice '...25RX
           SD10245Z', se prueba tambien cada pedazo de la descripcion).
        4. Si aun asi no hay match exacto, se prueba si algun pedazo es el
           final (sufijo) de un unico codigo del maestro -- algunos clientes
           escriben su propia OC con el codigo recortado (ej. '802X' en vez
           de 'RP9802X'). Si el sufijo hace match con varios codigos a la
           vez (ambiguo), se desempata comparando las palabras de la
           descripcion del cliente contra la descripcion de cada candidato
           en Product master list (ej. 'RESPIRATORY' en la OC prefiere el
           candidato cuya descripcion tambien dice 'Respiratory' sobre uno
           que diga 'GI-Bacteria'). Si sigue empatado, no se adivina.
        5. Si nada de lo anterior existe en el maestro (producto nuevo o
           Product master list desactualizado), pero el "codigo" que trajo
           el cliente es puramente numerico (o sea, probablemente solo el
           numero de renglon de SU PROPIA orden de compra, no un codigo real
           de Seegene -- los codigos de Seegene siempre traen letras) y la
           descripcion trae un codigo pegado despues de un '*' al final
           (ej. 'ANIPLEX II STI-7 *SD7700X'), se usa ese pedazo tal cual.
           Este ultimo caso NO esta validado contra Product master list (el
           diccionario resultante trae "from_master": False) -- se devuelve
           igual porque mostrar el codigo real de la OC (aunque no se pueda
           confirmar contra el maestro) es mas util que mostrar el numero de
           renglon del cliente, pero no se rellenan Descripcion/Categoria
           con datos inventados.
        El criterio final es simple: la primera parte (de donde sea) que
        exista de verdad en Product master list gana. Devuelve None si nada
        hace match -- nunca inventa un codigo validado contra el maestro."""
        if not raw_code and not description:
            return None
        candidates = self._tokenize(raw_code) + self._tokenize(description)
        # Alias conocidos (ver config.PRODUCT_CODE_ALIASES): codigo propio del
        # cliente/SAP que no se parece en nada al codigo real de Seegene, ya
        # confirmado a mano una vez -- se revisan primero para no depender de
        # que el tokenizado/sufijo adivine algo que nunca podria adivinar.
        for candidate in candidates:
            alias_target = PRODUCT_CODE_ALIASES.get(_norm(candidate))
            if alias_target:
                rec = self.product_by_code.get(_norm(alias_target))
                if rec:
                    rec = dict(rec)
                    rec["from_master"] = True
                    return rec
        for candidate in candidates:
            key = _norm(candidate)
            if key and key in self.product_by_code:
                rec = dict(self.product_by_code[key])
                rec["from_master"] = True
                return rec
        for candidate in candidates:
            key = _norm(candidate)
            # Se exige al menos un digito para evitar que una palabra generica
            # de la descripcion (ej. "SEEGENE", "TUBE") haga match por
            # casualidad con el final de un codigo real (ej. "173100-SEEGENE"
            # es un kit de herramientas, no tiene nada que ver con la palabra
            # "SEEGENE" que aparece en casi cualquier descripcion).
            if len(key) < 4 or not any(ch.isdigit() for ch in key):
                continue
            suffix_matches = [rec for code, rec in self.product_by_code.items() if code.endswith(key)]
            if len(suffix_matches) == 1:
                rec = dict(suffix_matches[0])
                rec["from_master"] = True
                return rec
            if len(suffix_matches) > 1 and description:
                desc_words = set(re.findall(r"[A-Za-z]+", str(description).upper()))
                scored = []
                for rec in suffix_matches:
                    rec_words = set(re.findall(r"[A-Za-z]+", str(rec.get("description") or "").upper()))
                    overlap = len(desc_words & rec_words)
                    if overlap:
                        scored.append((overlap, rec))
                if scored:
                    scored.sort(key=lambda x: -x[0])
                    if len(scored) == 1 or scored[0][0] > scored[1][0]:
                        rec = dict(scored[0][1])
                        rec["from_master"] = True
                        return rec
        if raw_code and re.fullmatch(r"\d+", str(raw_code).strip()) and description:
            m = re.search(r"\*\s*([A-Za-z][A-Za-z0-9\-]{2,})\s*$", str(description).strip())
            if m:
                parsed_code = m.group(1)
                key = _norm(parsed_code)
                result = {"code": parsed_code, "description": None, "category": None, "from_master": False}
                # A veces el codigo que trae la OC del cliente tiene un
                # pequeno error de captura frente al codigo real (ej.
                # 'SD7700X' impreso en la OC vs 'SD7701X' real en el
                # maestro) -- si hay un unico codigo en el maestro que
                # difiere en un solo caracter (misma longitud), se usa su
                # descripcion/categoria como referencia. El codigo mostrado
                # sigue siendo el que trae la OC del cliente (no se
                # reemplaza, porque no hay certeza de cual es el correcto).
                if key and key not in self.product_by_code:
                    near = [rec for code, rec in self.product_by_code.items()
                            if len(code) == len(key) and sum(a != b for a, b in zip(code, key)) == 1]
                    if len(near) == 1:
                        result["description"] = near[0]["description"]
                        result["category"] = near[0]["category"]
                return result
        return None
