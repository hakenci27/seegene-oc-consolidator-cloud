"""
extractor.py
Convierte cada archivo de Orden de Compra (PDF o Excel) en una lista de
partidas estructuradas, usando la API de Claude como "lector universal":
no importa si el documento es un PDF con texto, un PDF escaneado/manuscrito,
o un Excel con una plantilla distinta por cliente -- todos pasan por el
mismo extractor con visión, que es lo mismo que se hizo a mano durante el
desarrollo de este flujo.

Requiere: pip install anthropic pymupdf openpyxl
"""

import base64
import json
import re
from dataclasses import dataclass, field

import fitz  # PyMuPDF
import openpyxl
from anthropic import Anthropic

from config import CLAUDE_MODEL

EXTRACTION_PROMPT = """Eres un asistente que extrae datos de Ordenes de Compra (OC) \
de clientes de Seegene Mexico (laboratorios/hospitales/distribuidores que compran \
reactivos de diagnostico molecular).

Se te dara el contenido de UNA Orden de Compra (como imagen de cada pagina y/o texto \
plano extraido). Devuelve UNICAMENTE un JSON valido (sin explicaciones, sin markdown) \
con esta forma exacta:

{
  "po_number": "numero o folio de la orden de compra tal como aparece",
  "po_date": "fecha de la orden en formato YYYY-MM-DD, o null si no es clara",
  "customer_name": "nombre de la EMPRESA cliente/compradora (ver regla de abajo)",
  "customer_rfc": "RFC del cliente si aparece, o null",
  "notes": "cualquier ambiguedad relevante: fechas contradictorias, texto illegible, etc. o cadena vacia si no hay",
  "line_items": [
    {
      "code": "codigo/catalogo del producto tal como aparece (no lo inventes ni lo completes)",
      "description": "descripcion del producto",
      "qty": numero,
      "unit_price": numero,
      "code_uncertain": true/false  // true si el codigo esta manuscrito, incompleto o dudoso
    }
  ]
}

Reglas importantes:
- "customer_name" SIEMPRE debe ser el nombre de la EMPRESA (laboratorio/hospital/\
distribuidor) que compra, nunca el nombre de una persona. La forma mas confiable de \
identificarla es el logotipo o membrete de la empresa (usualmente en una esquina del \
encabezado) o un campo explicito tipo "Cliente"/"Empresa"/"Razon social". NO uses el \
nombre que aparece en campos como "Solicitante", "Elaboro", "Responsable", "Aprobo", \
"Reviso", "Atencion a" o "Contacto" -- esos son el nombre de UNA PERSONA (el empleado \
que llena el formulario), no el de la empresa. Tampoco uses el campo "Proveedor" (ese \
es Seegene, el vendedor, nunca el cliente).
- Si un renglon de la tabla tiene cantidad 0 o esta vacio, NO lo incluyas.
- Si el documento esta escaneado o manuscrito y algun codigo de producto no se lee con \
certeza, transcribe tu mejor lectura y marca "code_uncertain": true -- nunca inventes un \
codigo que no puedas justificar con lo que ves.
- Si hay dos fechas distintas e inconsistentes en el documento, usa la que parezca mas \
confiable y explica la discrepancia en "notes".
- Responde SOLO con el JSON, nada de texto antes o despues.
"""


@dataclass
class LineItem:
    code: str
    description: str
    qty: float
    unit_price: float
    code_uncertain: bool = False


@dataclass
class ExtractedPO:
    source_file: str
    po_number: str
    po_date: str | None
    customer_name: str
    customer_rfc: str | None
    notes: str
    line_items: list = field(default_factory=list)


def _pdf_to_images_and_text(pdf_path, max_pages=5, zoom=2.0):
    """Renderiza cada pagina de un PDF a PNG (base64) y tambien intenta sacar texto."""
    doc = fitz.open(pdf_path)
    images_b64 = []
    text_parts = []
    for i, page in enumerate(doc):
        if i >= max_pages:
            break
        text_parts.append(page.get_text())
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        png_bytes = pix.tobytes("png")
        images_b64.append(base64.b64encode(png_bytes).decode("ascii"))
    doc.close()
    return images_b64, "\n".join(text_parts).strip()


def _xlsx_to_text(xlsx_path, max_rows=200):
    """Convierte todas las hojas de un Excel a texto plano tipo CSV para dárselo a Claude."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chunks.append(f"--- Hoja: {sheet_name} ---")
        for r, row in enumerate(ws.iter_rows(values_only=True)):
            if r >= max_rows:
                break
            if any(v is not None for v in row):
                chunks.append(" | ".join("" if v is None else str(v) for v in row))
    return "\n".join(chunks)


def _call_claude(client, images_b64, text_content, source_label):
    content = []
    if text_content:
        content.append({
            "type": "text",
            "text": f"Texto extraido automaticamente del documento (puede estar incompleto):\n{text_content}\n"
        })
    for img_b64 in images_b64:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": img_b64},
        })
    content.append({"type": "text", "text": EXTRACTION_PROMPT})

    message = client.messages.create(
        model=CLAUDE_MODEL,
        max_tokens=4096,
        messages=[{"role": "user", "content": content}],
    )
    raw = "".join(block.text for block in message.content if block.type == "text")
    raw = raw.strip()
    # por si Claude envuelve el JSON en ```json ... ```
    match = re.search(r"\{.*\}", raw, re.DOTALL)
    if not match:
        raise ValueError(f"No se pudo interpretar la respuesta de Claude para {source_label}:\n{raw[:500]}")
    data = json.loads(match.group(0))
    return data


def extract_po(file_path, api_key, log=print):
    """
    Extrae una Orden de Compra (PDF o xlsx) usando la API de Claude.
    Devuelve un ExtractedPO. Lanza excepcion si algo sale mal (el llamador
    decide como reportarlo en la interfaz).
    """
    client = Anthropic(api_key=api_key)
    lower = file_path.lower()

    if lower.endswith(".pdf"):
        log(f"  Leyendo PDF: {file_path}")
        images_b64, text_content = _pdf_to_images_and_text(file_path)
        data = _call_claude(client, images_b64, text_content, file_path)
    elif lower.endswith((".xlsx", ".xlsm")):
        log(f"  Leyendo Excel: {file_path}")
        text_content = _xlsx_to_text(file_path)
        data = _call_claude(client, [], text_content, file_path)
    else:
        raise ValueError(f"Tipo de archivo no soportado: {file_path}")

    items = [
        LineItem(
            code=str(li.get("code", "")).strip(),
            description=str(li.get("description", "")).strip(),
            qty=float(li.get("qty") or 0),
            unit_price=float(li.get("unit_price") or 0),
            code_uncertain=bool(li.get("code_uncertain", False)),
        )
        for li in data.get("line_items", [])
        if float(li.get("qty") or 0) > 0
    ]

    return ExtractedPO(
        source_file=file_path,
        po_number=str(data.get("po_number", "")).strip(),
        po_date=data.get("po_date"),
        customer_name=str(data.get("customer_name", "")).strip(),
        customer_rfc=(data.get("customer_rfc") or None),
        notes=str(data.get("notes", "")).strip(),
        line_items=items,
    )
